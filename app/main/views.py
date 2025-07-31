import os
import uuid
import shutil
import json
from datetime import datetime
from io import BytesIO
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from flask import render_template, redirect, url_for, flash, request, current_app, jsonify, abort, send_from_directory, send_file
from flask_login import login_required, current_user
from .. import csrf
from . import main
from .. import db
from ..models import Order, OrderImage, Permission, OrderField, OrderType, WechatUser, User
from ..forms import OrderForm
from ..decorators import admin_required
from werkzeug.utils import secure_filename

def allowed_file(filename):
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in current_app.config['ALLOWED_EXTENSIONS']

def save_image(file, subfolder=''):
    if file and allowed_file(file.filename):
        # 检查文件大小
        file.seek(0, 2)  # 移动到文件末尾
        file_size = file.tell()
        file.seek(0)  # 重置文件指针
        
        max_size = current_app.config.get('MAX_CONTENT_LENGTH', 16 * 1024 * 1024)  # 默认16MB
        if file_size > max_size:
            raise ValueError(f"文件大小超过限制 ({max_size // (1024*1024)}MB)")
        
        filename = secure_filename(file.filename)
        # 获取文件扩展名
        file_ext = filename.rsplit('.', 1)[1].lower() if '.' in filename else ''
        
        # 验证文件类型（通过魔数检查）
        file_header = file.read(8)
        file.seek(0)
        
        # 简单的文件类型验证
        image_signatures = {
            b'\xff\xd8\xff': 'jpg',
            b'\x89PNG\r\n\x1a\n': 'png',
            b'GIF87a': 'gif',
            b'GIF89a': 'gif',
            b'RIFF': 'webp'  # WebP文件以RIFF开头
        }
        
        is_valid_image = False
        for signature in image_signatures:
            if file_header.startswith(signature):
                is_valid_image = True
                break
        
        if not is_valid_image:
            raise ValueError("无效的图片文件格式")
        
        # 生成唯一文件名，确保包含扩展名
        unique_filename = f"{uuid.uuid4().hex}.{file_ext}"
        
        # 创建子目录路径
        if subfolder:
            upload_dir = os.path.join(current_app.config['UPLOAD_FOLDER'], subfolder)
            os.makedirs(upload_dir, exist_ok=True)
            file_path = os.path.join(upload_dir, unique_filename)
            file.save(file_path)
            # 返回相对路径，用于存储在数据库（使用正斜杠以确保Web兼容性）
            return f"{subfolder}/{unique_filename}"
        else:
            file_path = os.path.join(current_app.config['UPLOAD_FOLDER'], unique_filename)
            file.save(file_path)
            # 返回相对路径，用于存储在数据库（使用正斜杠以确保Web兼容性）
            return f"{unique_filename}"
    return None

@main.route('/')
def index():
    if current_user.is_authenticated:
        return redirect(url_for('main.order_list'))
    return render_template('index.html')

@main.route('/orders')
@login_required
def order_list():
    from datetime import date, timedelta
    from sqlalchemy import and_, extract
    
    page = request.args.get('page', 1, type=int)
    user_id = request.args.get('user_id', type=int)
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    search_type = request.args.get('search_type', 'wechat_name')
    search_value = request.args.get('search_value', '').strip()
    sort_by = request.args.get('sort_by')  # 排序方式：amount(金额)、count(数量) 或 None(按时间)
    
    # 构建查询
    query = Order.query
    
    # 权限控制
    if current_user.can(Permission.VIEW_ALL):
        # 管理员可以查看所有订单，可以按用户筛选
        if user_id:
            query = query.filter(Order.user_id == user_id)
    else:
        # 普通用户只能查看自己的订单
        query = query.filter(Order.user_id == current_user.id)
    
    # 日期筛选，默认当前月份
    if not start_date and not end_date:
        # 默认显示当前月份
        today = date.today()
        start_date = today.replace(day=1).strftime('%Y-%m-%d')
        # 计算月末
        if today.month == 12:
            next_month = today.replace(year=today.year + 1, month=1, day=1)
        else:
            next_month = today.replace(month=today.month + 1, day=1)
        end_date = (next_month - timedelta(days=1)).strftime('%Y-%m-%d')
    
    if start_date:
        try:
            start_dt = datetime.strptime(start_date, '%Y-%m-%d')
            query = query.filter(Order.completion_time >= start_dt)
        except ValueError:
            flash('开始日期格式错误', 'danger')
    
    if end_date:
        try:
            end_dt = datetime.strptime(end_date, '%Y-%m-%d')
            query = query.filter(Order.completion_time <= end_dt)
        except ValueError:
            flash('结束日期格式错误', 'danger')
    
    # 搜索功能
    if search_value:
        if search_type == 'wechat_name':
            query = query.filter(Order.wechat_name.like(f'%{search_value}%'))
        elif search_type == 'wechat_id':
            query = query.filter(Order.wechat_id.like(f'%{search_value}%'))
        elif search_type == 'phone':
            query = query.filter(Order.phone.like(f'%{search_value}%'))
        elif search_type == 'order_code':
            query = query.filter(Order.order_code.like(f'%{search_value}%'))
    
    # 排序 - 默认按创建时间降序，新订单在前
    if sort_by == 'amount':
        query = query.order_by(Order.amount.desc().nullslast(), Order.create_time.desc())
    elif sort_by == 'count':
        query = query.order_by(Order.quantity.desc().nullslast(), Order.create_time.desc())
    else:
        # 默认排序：按创建时间降序，确保新订单在最前面
        query = query.order_by(Order.create_time.desc())
    
    # 分页
    pagination = query.paginate(
        page=page, per_page=10, error_out=False
    )
    orders = pagination.items
    
    # 获取所有用户（用于筛选）
    users = []
    if current_user.can(Permission.VIEW_ALL):
        from ..models import User
        users = User.query.all()
    
    # 计算统计信息
    from sqlalchemy import func
    total_orders = pagination.total
    total_amount = query.with_entities(func.sum(Order.amount)).scalar() or 0
    avg_amount = query.with_entities(func.avg(Order.amount)).scalar() or 0
    total_quantity = query.with_entities(func.sum(Order.quantity)).scalar() or 0
    
    # 计算微信用户数（去重）
    total_wechat_users = query.with_entities(func.count(func.distinct(Order.wechat_name))).scalar() or 0
    
    # 构建当前筛选条件
    current_filters = {
        'user_id': user_id,
        'start_date': start_date,
        'end_date': end_date,
        'search_type': search_type,
        'search_value': search_value,
        'sort_by': sort_by
    }
    
    return render_template('main/order_list.html',
                         orders=orders,
                         pagination=pagination,
                         users=users,
                         current_filters=current_filters,
                         total_orders=total_orders,
                         total_amount=total_amount,
                         avg_amount=avg_amount,
                         total_quantity=total_quantity,
                         total_wechat_users=total_wechat_users,
                         now=datetime.now())

@main.route('/order/new', methods=['GET', 'POST'])
@login_required
def new_order():
    form = OrderForm()
    
    if form.validate_on_submit():
        # 生成订单编号
        order_code = f"ORD{datetime.now().strftime('%Y%m%d%H%M%S')}{uuid.uuid4().hex[:8].upper()}"
        
        # 创建订单
        order = Order(
            order_code=order_code,
            wechat_name=form.wechat_name.data,
            wechat_id=form.wechat_id.data,
            phone=form.phone.data,
            order_info=form.order_info.data,
            completion_time=form.completion_time.data,
            quantity=form.quantity.data,
            amount=form.amount.data,
            notes=form.notes.data,
            user_id=current_user.id,
            order_type_id=form.order_type_id.data
        )
        
        # 处理自定义字段
        custom_fields = {}
        for field in OrderField.query.filter_by(is_default=False).all():
            if hasattr(form, field.name):
                field_value = getattr(form, field.name).data
                if field_value:
                    custom_fields[field.name] = field_value
        order.custom_fields = json.dumps(custom_fields, ensure_ascii=False) if custom_fields else None
        
        # 处理图片上传
        uploaded_files = request.files.getlist('images')
        for uploaded_file in uploaded_files:
            if uploaded_file and uploaded_file.filename:
                image_path = save_image(uploaded_file, 'orders')
                if image_path:
                    order_image = OrderImage(order_id=order.id, image_path=image_path)
                    db.session.add(order_image)
        
        db.session.add(order)
        db.session.commit()
        
        flash('订单创建成功！', 'success')
        return redirect(url_for('main.order_list'))
    
    # GET请求或表单验证失败，显示表单
    return render_template('main/new_order.html', form=form)

@main.route('/order/<int:id>')
@login_required
def view_order(id):
    order = Order.query.get_or_404(id)
    
    # 权限检查
    if not current_user.can(Permission.VIEW_ALL) and order.user_id != current_user.id:
        abort(403)
    
    return render_template('main/view_order.html', order=order)

@main.route('/order/edit/<int:id>', methods=['GET', 'POST'])
@login_required
def edit_order(id):
    order = Order.query.get_or_404(id)
    
    # 权限检查
    if not current_user.can(Permission.VIEW_ALL) and order.user_id != current_user.id:
        abort(403)
    
    form = OrderForm(obj=order)
    form.order = order  # 设置当前订单对象，用于验证时排除自身
    form.order_type_id.choices = [(t.id, t.name) for t in OrderType.query.filter_by(is_active=True).all()]
    
    if form.validate_on_submit():
        # 更新订单信息
        order.wechat_name = form.wechat_name.data
        order.wechat_id = form.wechat_id.data
        order.phone = form.phone.data
        order.order_info = form.order_info.data
        order.quantity = form.quantity.data
        order.amount = form.amount.data
        order.notes = form.notes.data
        order.order_type_id = form.order_type_id.data
        order.completion_time = form.completion_time.data
        
        # 处理自定义字段
        custom_fields = {}
        for field in OrderField.query.filter_by(is_default=False).all():
            if hasattr(form, f'custom_{field.name}'):
                field_value = getattr(form, f'custom_{field.name}').data
                if field_value:
                    custom_fields[field.name] = field_value
        order.custom_fields = json.dumps(custom_fields, ensure_ascii=False) if custom_fields else None
        
        # 处理图片上传
        uploaded_files = request.files.getlist('images')
        for uploaded_file in uploaded_files:
            if uploaded_file and uploaded_file.filename:
                image_path = save_image(uploaded_file, 'orders')
                if image_path:
                    order_image = OrderImage(order_id=order.id, image_path=image_path)
                    db.session.add(order_image)
        
        db.session.commit()
        flash('订单更新成功！', 'success')
        return redirect(url_for('main.view_order', id=id))
    
    # GET请求或表单验证失败，显示编辑表单
    return render_template('main/edit_order.html', form=form, order=order)

@main.route('/order/delete/<int:id>', methods=['POST'])
@login_required
def delete_order(id):
    order = Order.query.get_or_404(id)
    
    # 权限检查
    if not current_user.can(Permission.VIEW_ALL) and order.user_id != current_user.id:
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({'success': False, 'error': '权限不足'})
        abort(403)
    
    try:
        # 删除相关图片
        for image in order.images:
            try:
                image_path = os.path.join(current_app.config['UPLOAD_FOLDER'], image.image_path.replace('uploads/', ''))
                if os.path.exists(image_path):
                    os.remove(image_path)
            except Exception as e:
                print(f"删除图片失败: {e}")
                pass
            db.session.delete(image)
        
        db.session.delete(order)
        db.session.commit()
        
        # 根据请求类型返回不同响应
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({'success': True, 'message': '订单删除成功！'})
        else:
            flash('订单删除成功！', 'success')
            return redirect(url_for('main.order_list'))
            
    except Exception as e:
        db.session.rollback()
        print(f"删除订单失败: {e}")
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({'success': False, 'error': f'删除失败：{str(e)}'})
        else:
            flash(f'删除失败：{str(e)}', 'error')
            return redirect(url_for('main.order_list'))

@main.route('/orders/statistics')
@login_required
def order_statistics():
    # 仅超级管理员可访问
    if not current_user.can(Permission.VIEW_ALL):
        abort(403)
    
    from datetime import date, timedelta
    from sqlalchemy import func
    
    # 获取查询参数
    user_id = request.args.get('user_id', type=int)
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    search_type = request.args.get('search_type', 'wechat_name')
    search_value = request.args.get('search_value', '').strip()
    sort_by = request.args.get('sort_by', 'amount')
    
    # 默认显示所有数据（不设置日期范围）
    # 如果用户没有指定日期范围，则显示所有数据
    if not start_date and not end_date:
        # 不设置默认日期范围，显示所有数据
        pass
    
    # 构建查询
    query = Order.query
    start_dt = None
    end_dt = None
    
    # 用户筛选
    if user_id:
        query = query.filter(Order.user_id == user_id)
    
    # 日期筛选
    if start_date:
        try:
            start_dt = datetime.strptime(start_date, '%Y-%m-%d')
            query = query.filter(Order.create_time >= start_dt)
        except ValueError:
            flash('开始日期格式错误', 'danger')
    
    if end_date:
        try:
            end_dt = datetime.strptime(end_date, '%Y-%m-%d')
            # 结束日期包含当天，所以加1天
            from datetime import timedelta
            end_dt = end_dt + timedelta(days=1)
            query = query.filter(Order.create_time < end_dt)
        except ValueError:
            flash('结束日期格式错误', 'danger')
    
    # 搜索筛选
    if search_value:
        if search_type == 'wechat_name':
            query = query.filter(Order.wechat_name.contains(search_value))
        elif search_type == 'wechat_id':
            query = query.filter(Order.wechat_id.contains(search_value))
        elif search_type == 'phone':
            query = query.filter(Order.phone.contains(search_value))
    
    # 统计信息
    total_orders = query.count()
    total_amount = query.with_entities(func.sum(Order.amount)).scalar() or 0
    avg_amount = query.with_entities(func.avg(Order.amount)).scalar() or 0
    total_quantity = query.with_entities(func.sum(Order.quantity)).scalar() or 0
    
    # 按状态统计
    status_stats_query = db.session.query(
        Order.status,
        func.count(Order.id).label('count'),
        func.sum(Order.amount).label('amount')
    )
    
    # 应用筛选条件
    if user_id:
        status_stats_query = status_stats_query.filter(Order.user_id == user_id)
    if start_dt:
        status_stats_query = status_stats_query.filter(Order.create_time >= start_dt)
    if end_dt:
        status_stats_query = status_stats_query.filter(Order.create_time < end_dt)
    if search_value:
        if search_type == 'wechat_name':
            status_stats_query = status_stats_query.filter(Order.wechat_name.contains(search_value))
        elif search_type == 'wechat_id':
            status_stats_query = status_stats_query.filter(Order.wechat_id.contains(search_value))
        elif search_type == 'phone':
            status_stats_query = status_stats_query.filter(Order.phone.contains(search_value))
    
    status_stats = status_stats_query.group_by(Order.status).all()
    
    # 按订单类型统计
    type_stats_query = db.session.query(
        OrderType.name,
        func.count(Order.id).label('order_count'),
        func.sum(Order.amount).label('total_amount')
    ).outerjoin(Order, OrderType.id == Order.order_type_id)
    
    # 应用筛选条件
    if user_id:
        type_stats_query = type_stats_query.filter(Order.user_id == user_id)
    if start_dt:
        type_stats_query = type_stats_query.filter(Order.create_time >= start_dt)
    if end_dt:
        type_stats_query = type_stats_query.filter(Order.create_time < end_dt)
    if search_value:
        if search_type == 'wechat_name':
            type_stats_query = type_stats_query.filter(Order.wechat_name.contains(search_value))
        elif search_type == 'wechat_id':
            type_stats_query = type_stats_query.filter(Order.wechat_id.contains(search_value))
        elif search_type == 'phone':
            type_stats_query = type_stats_query.filter(Order.phone.contains(search_value))
    
    type_stats = type_stats_query.group_by(OrderType.id, OrderType.name).all()
    
    # 按用户统计
    user_stats_query = db.session.query(
        User.username,
        func.count(Order.id).label('order_count'),
        func.sum(Order.amount).label('total_amount')
    ).join(Order, User.id == Order.user_id)
    
    # 应用筛选条件
    if user_id:
        user_stats_query = user_stats_query.filter(Order.user_id == user_id)
    if start_dt:
        user_stats_query = user_stats_query.filter(Order.create_time >= start_dt)
    if end_dt:
        user_stats_query = user_stats_query.filter(Order.create_time < end_dt)
    if search_value:
        if search_type == 'wechat_name':
            user_stats_query = user_stats_query.filter(Order.wechat_name.contains(search_value))
        elif search_type == 'wechat_id':
            user_stats_query = user_stats_query.filter(Order.wechat_id.contains(search_value))
        elif search_type == 'phone':
            user_stats_query = user_stats_query.filter(Order.phone.contains(search_value))
    
    user_stats = user_stats_query.group_by(User.id, User.username).all()
    
    # 按微信用户统计（Top 10）
    wechat_stats_query = db.session.query(
        Order.wechat_name,
        func.count(Order.id).label('order_count'),
        func.sum(Order.amount).label('total_amount')
    )
    
    # 应用筛选条件
    if user_id:
        wechat_stats_query = wechat_stats_query.filter(Order.user_id == user_id)
    if start_dt:
        wechat_stats_query = wechat_stats_query.filter(Order.create_time >= start_dt)
    if end_dt:
        wechat_stats_query = wechat_stats_query.filter(Order.create_time < end_dt)
    if search_value:
        if search_type == 'wechat_name':
            wechat_stats_query = wechat_stats_query.filter(Order.wechat_name.contains(search_value))
        elif search_type == 'wechat_id':
            wechat_stats_query = wechat_stats_query.filter(Order.wechat_id.contains(search_value))
        elif search_type == 'phone':
            wechat_stats_query = wechat_stats_query.filter(Order.phone.contains(search_value))
    
    wechat_stats_query = wechat_stats_query.filter(Order.wechat_name.isnot(None)).group_by(Order.wechat_name)
    
    if sort_by == 'count':
        wechat_stats = wechat_stats_query.order_by(func.count(Order.id).desc()).limit(10).all()
    else:
        wechat_stats = wechat_stats_query.order_by(func.sum(Order.amount).desc()).limit(10).all()
    
    # 获取所有用户（用于筛选）
    users = User.query.all()
    
    # 构建当前筛选条件
    current_filters = {
        'user_id': user_id,
        'start_date': start_date,
        'end_date': end_date,
        'search_type': search_type,
        'search_value': search_value,
        'sort_by': sort_by
    }
    
    return render_template('main/order_statistics.html',
                         total_orders=total_orders,
                         total_amount=total_amount,
                         avg_amount=avg_amount,
                         total_quantity=total_quantity,
                         status_stats=status_stats,
                         type_stats=type_stats,
                         user_stats=user_stats,
                         wechat_stats=wechat_stats,
                         users=users,
                         current_filters=current_filters)

@main.route('/debug/user-info')
@login_required
def debug_user_info():
    """调试用户信息"""
    if not current_user.is_administrator():
        abort(403)
    
    return jsonify({
        'user_id': current_user.id,
        'username': current_user.username,
        'email': current_user.email,
        'role': current_user.role.name if current_user.role else None,
        'permissions': {
            'VIEW_OWN': current_user.can(Permission.VIEW_OWN),
            'SUBMIT': current_user.can(Permission.SUBMIT),
            'VIEW_ALL': current_user.can(Permission.VIEW_ALL),
            'MANAGE_FIELDS': current_user.can(Permission.MANAGE_FIELDS),
            'ADMIN': current_user.can(Permission.ADMIN)
        }
    })

@main.route('/orders/export-template')
@login_required
def export_template():
    """导出订单模板"""
    # 所有登录用户都可以下载模板
    # 移除权限检查，因为下载模板是基础功能
    
    # 导入语句已在文件顶部
    
    # 创建工作簿
    wb = Workbook()
    
    # 创建订单模板sheet
    ws1 = wb.active
    ws1.title = "订单模板"
    
    # 获取当前日期作为默认完成时间
    current_date = datetime.now().strftime('%Y-%m-%d')
    
    # 模板数据 - 按页面列顺序排列，备注在最后
    template_data = {
        '*订单编码': ['ORD20240101001'],
        '*订单类型': ['海报'],
        '*微信名': ['张三'],
        '*手机号': ['13800138001'],
        '*订单信息': ['制作海报，尺寸A4'],
        '*完成时间': [current_date],
        '*数量': [1],
        '金额': [100.00],
        '备注': ['']
    }
    
    df = pd.DataFrame(template_data)
    
    # 将数据写入工作表
    for r in dataframe_to_rows(df, index=False, header=True):
        ws1.append(r)
    
    # 手动设置带星号的列标题，使星号为红色
    from openpyxl.styles.colors import Color
    from openpyxl.cell.rich_text import TextBlock, CellRichText
    from openpyxl.cell.text import InlineFont
    
    # 设置必填项列标题的星号为红色
    red_star = TextBlock(InlineFont(b=True, color="FF0000", sz=12), "*")
    
    # A列：*订单编码
    white_text_a = TextBlock(InlineFont(b=True, color="FFFFFF", sz=12), "订单编码")
    ws1['A1'].value = CellRichText(red_star, white_text_a)
    
    # B列：*订单类型
    white_text_b = TextBlock(InlineFont(b=True, color="FFFFFF", sz=12), "订单类型")
    ws1['B1'].value = CellRichText(red_star, white_text_b)
    
    # C列：*微信名
    white_text_c = TextBlock(InlineFont(b=True, color="FFFFFF", sz=12), "微信名")
    ws1['C1'].value = CellRichText(red_star, white_text_c)
    
    # D列：*手机号
    white_text_d = TextBlock(InlineFont(b=True, color="FFFFFF", sz=12), "手机号")
    ws1['D1'].value = CellRichText(red_star, white_text_d)
    
    # E列：*订单信息
    white_text_e = TextBlock(InlineFont(b=True, color="FFFFFF", sz=12), "订单信息")
    ws1['E1'].value = CellRichText(red_star, white_text_e)
    
    # F列：*完成时间
    white_text_f = TextBlock(InlineFont(b=True, color="FFFFFF", sz=12), "完成时间")
    ws1['F1'].value = CellRichText(red_star, white_text_f)
    
    # G列：*数量
    white_text_g = TextBlock(InlineFont(b=True, color="FFFFFF", sz=12), "数量")
    ws1['G1'].value = CellRichText(red_star, white_text_g)
    
    # 设置丰富的样式
    header_font = Font(bold=True, color="FFFFFF", size=12, name="微软雅黑")
    header_fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # 数据行样式
    data_font = Font(size=11, name="微软雅黑")
    data_fill_light = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
    data_fill_white = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    data_alignment = Alignment(horizontal="center", vertical="center")  # 内容居中
    number_alignment = Alignment(horizontal="center", vertical="center")  # 数字也居中
    
    # 设置边框
    thin_border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )
    
    thick_border = Border(
        left=Side(style='medium', color='2F5597'),
        right=Side(style='medium', color='2F5597'),
        top=Side(style='medium', color='2F5597'),
        bottom=Side(style='medium', color='2F5597')
    )
    
    # 应用标题行样式
    for col_idx, cell in enumerate(ws1[1], 1):
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thick_border
        cell.font = header_font  # 统一使用白色字体
    
    # 设置列宽
    column_widths = {
        'A': 18,  # *订单编码
        'B': 12,  # *订单类型
        'C': 15,  # *微信名
        'D': 15,  # *手机号
        'E': 25,  # *订单信息
        'F': 12,  # *完成时间
        'G': 8,   # *数量
        'H': 10,  # 金额
        'I': 20   # 备注
    }
    
    for col, width in column_widths.items():
        ws1.column_dimensions[col].width = width
    
    # 为所有数据单元格添加边框和样式
    for row_idx, row in enumerate(ws1.iter_rows(min_row=1, max_row=ws1.max_row, min_col=1, max_col=ws1.max_column), 1):
        for col_idx, cell in enumerate(row, 1):
            cell.border = thin_border
            if cell.row > 1:  # 非标题行
                cell.font = data_font
                # 交替行颜色
                if cell.row % 2 == 0:
                    cell.fill = data_fill_light
                else:
                    cell.fill = data_fill_white
                
                # 所有内容都居中对齐
                cell.alignment = data_alignment
    
    # 创建字段说明sheet
    ws2 = wb.create_sheet(title="字段说明")
    
    instructions_data = {
        '字段名': ['订单编号', '订单类型', '*微信名', '*手机号', '完成时间', '数量', '金额', '备注'],
        '说明': [
            '系统自动生成，无需填写',
            '订单类型：海报/名片/宣传册等',
            '客户微信昵称（必填，标*为红色）',
            '客户手机号（必填，标*为红色）',
            '订单完成时间（YYYY-MM-DD格式，默认当前日期）',
            '商品数量（整数）',
            '订单金额（数字）',
            '备注信息'
        ],
        '是否必填': ['否', '否', '是', '是', '否', '否', '否', '否'],
        '数据类型': ['文本', '文本', '文本', '文本', '日期', '数字', '数字', '文本']
    }
    
    instructions_df = pd.DataFrame(instructions_data)
    
    # 将说明数据写入工作表
    for r in dataframe_to_rows(instructions_df, index=False, header=True):
        ws2.append(r)
    
    # 应用说明sheet样式
    for cell in ws2[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thick_border
    
    # 设置说明sheet列宽
    ws2.column_dimensions['A'].width = 15
    ws2.column_dimensions['B'].width = 45
    ws2.column_dimensions['C'].width = 12
    ws2.column_dimensions['D'].width = 12
    
    # 为说明sheet所有单元格添加边框和样式
    for row_idx, row in enumerate(ws2.iter_rows(min_row=1, max_row=ws2.max_row, min_col=1, max_col=ws2.max_column), 1):
        for col_idx, cell in enumerate(row, 1):
            cell.border = thin_border
            if cell.row > 1:
                cell.font = data_font
                # 交替行颜色
                if cell.row % 2 == 0:
                    cell.fill = data_fill_light
                else:
                    cell.fill = data_fill_white
                
                # 设置对齐和换行
                if col_idx == 2:  # 说明列
                    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                else:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # 保存到BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'订单导入模板_{datetime.now().strftime("%Y%m%d")}.xlsx'
    )

@main.route('/orders/export')
@login_required
def export_orders():
    """导出订单数据"""
    # 所有登录用户都可以导出订单
    # 普通用户只能导出自己的订单，管理员可以导出所有订单
    
    import pandas as pd
    from io import BytesIO
    
    # 获取查询参数
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    user_id = request.args.get('user_id', type=int)
    
    # 构建查询
    query = Order.query
    
    # 权限控制：普通用户只能导出自己的订单
    if current_user.can(Permission.VIEW_ALL):
        # 管理员可以按用户筛选
        if user_id:
            query = query.filter(Order.user_id == user_id)
    else:
        # 普通用户只能导出自己的订单
        query = query.filter(Order.user_id == current_user.id)
    
    if start_date:
        try:
            start_dt = datetime.strptime(start_date, '%Y-%m-%d')
            query = query.filter(Order.completion_time >= start_dt)
        except ValueError:
            flash('开始日期格式错误', 'danger')
    
    if end_date:
        try:
            end_dt = datetime.strptime(end_date, '%Y-%m-%d')
            query = query.filter(Order.completion_time <= end_dt)
        except ValueError:
            flash('结束日期格式错误', 'danger')
    
    orders = query.order_by(Order.create_time.desc()).all()
    
    # 准备导出数据
    export_data = []
    for order in orders:
        # 解析自定义字段
        custom_fields = {}
        if order.custom_fields:
            try:
                custom_fields = json.loads(order.custom_fields)
            except:
                custom_fields = {}
        
        row = {
            '订单编号': order.order_code,
            '微信名': order.wechat_name,
            '微信号': order.wechat_id,
            '手机号': order.phone,
            '订单信息': order.order_info,
            '完成时间': order.completion_time.strftime('%Y-%m-%d') if order.completion_time else '',
            '数量': order.quantity,
            '金额': order.amount,
            '备注': order.notes,
            '订单类型': order.order_type.name if order.order_type else '',
            '状态': order.status,
            '创建时间': order.create_time.strftime('%Y-%m-%d %H:%M:%S'),
            '创建用户': order.creator.username if order.creator else ''
        }
        
        # 添加自定义字段
        for field_name, field_value in custom_fields.items():
            row[f'自定义_{field_name}'] = field_value
        
        export_data.append(row)
    
    if not export_data:
        flash('没有找到符合条件的订单', 'warning')
        return redirect(url_for('main.order_list'))
    
    df = pd.DataFrame(export_data)
    
    # 创建工作簿
    wb = Workbook()
    
    # 创建订单数据sheet
    ws1 = wb.active
    ws1.title = "订单数据"
    
    # 将数据写入工作表
    for r in dataframe_to_rows(df, index=False, header=True):
        ws1.append(r)
    
    # 设置丰富的样式
    header_font = Font(bold=True, color="FFFFFF", size=12, name="微软雅黑")
    header_fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # 数据行样式
    data_font = Font(size=10, name="微软雅黑")
    data_fill_light = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
    data_fill_white = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    data_alignment = Alignment(horizontal="left", vertical="center")
    number_alignment = Alignment(horizontal="right", vertical="center")
    date_alignment = Alignment(horizontal="center", vertical="center")
    
    # 设置边框
    thin_border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )
    
    thick_border = Border(
        left=Side(style='medium', color='2F5597'),
        right=Side(style='medium', color='2F5597'),
        top=Side(style='medium', color='2F5597'),
        bottom=Side(style='medium', color='2F5597')
    )
    
    # 应用标题行样式
    for cell in ws1[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thick_border
    
    # 设置列宽
    column_widths = {
        'A': 18,  # 订单编号
        'B': 12,  # 微信名
        'C': 15,  # 微信号
        'D': 15,  # 手机号
        'E': 25,  # 订单信息
        'F': 12,  # 完成时间
        'G': 8,   # 数量
        'H': 12,  # 金额
        'I': 20,  # 备注
        'J': 12,  # 订单类型
        'K': 10,  # 状态
        'L': 18,  # 创建时间
        'M': 12   # 创建用户
    }
    
    for col, width in column_widths.items():
        if col in [chr(65+i) for i in range(ws1.max_column)]:
            ws1.column_dimensions[col].width = width
    
    # 为所有数据单元格添加边框和样式
    for row_idx, row in enumerate(ws1.iter_rows(min_row=1, max_row=ws1.max_row, min_col=1, max_col=ws1.max_column), 1):
        for col_idx, cell in enumerate(row, 1):
            cell.border = thin_border
            if cell.row > 1:  # 非标题行
                cell.font = data_font
                # 交替行颜色
                if cell.row % 2 == 0:
                    cell.fill = data_fill_light
                else:
                    cell.fill = data_fill_white
                
                # 根据列类型设置对齐方式
                if col_idx in [7, 8]:  # 数量和金额列
                    cell.alignment = number_alignment
                elif col_idx in [6, 12]:  # 完成时间和创建时间列
                    cell.alignment = date_alignment
                else:
                    cell.alignment = data_alignment
    
    # 创建统计信息sheet
    ws2 = wb.create_sheet(title="统计信息")
    
    # 添加统计信息
    stats_data = {
        '统计项': ['总订单数', '总金额', '平均金额', '总数量'],
        '数值': [
            len(export_data),
            sum(row['金额'] for row in export_data if row['金额']),
            round(sum(row['金额'] for row in export_data if row['金额']) / len(export_data), 2) if export_data else 0,
            sum(row['数量'] for row in export_data if row['数量'])
        ]
    }
    
    stats_df = pd.DataFrame(stats_data)
    
    # 将统计数据写入工作表
    for r in dataframe_to_rows(stats_df, index=False, header=True):
        ws2.append(r)
    
    # 应用统计sheet样式
    for cell in ws2[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thick_border
    
    # 设置统计sheet列宽
    ws2.column_dimensions['A'].width = 15
    ws2.column_dimensions['B'].width = 15
    
    # 为统计sheet所有单元格添加边框和样式
    for row_idx, row in enumerate(ws2.iter_rows(min_row=1, max_row=ws2.max_row, min_col=1, max_col=ws2.max_column), 1):
        for col_idx, cell in enumerate(row, 1):
            cell.border = thin_border
            if cell.row > 1:
                cell.font = data_font
                # 交替行颜色
                if cell.row % 2 == 0:
                    cell.fill = data_fill_light
                else:
                    cell.fill = data_fill_white
                
                # 数值列右对齐
                if col_idx == 2:
                    cell.alignment = number_alignment
                else:
                    cell.alignment = data_alignment
    
    # 保存到BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'订单导出_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    )

@main.route('/orders/import', methods=['GET', 'POST'])
@login_required
def import_orders():
    """导入订单数据"""
    # 所有登录用户都可以导入订单
    # 普通用户导入的订单将归属于自己
    
    if request.method == 'POST':
        if 'file' not in request.files:
            flash('没有选择文件', 'danger')
            return redirect(request.url)
        
        file = request.files['file']
        if file.filename == '':
            flash('没有选择文件', 'danger')
            return redirect(request.url)
        
        if file and allowed_file(file.filename):
            try:
                import pandas as pd
                
                # 检查文件扩展名
                file_ext = file.filename.rsplit('.', 1)[1].lower() if '.' in file.filename else ''
                
                # 读取文件
                if file_ext in ['xlsx', 'xls']:
                    df = pd.read_excel(file)
                elif file_ext == 'csv':
                    df = pd.read_csv(file, encoding='utf-8')
                else:
                    flash(f'不支持的文件格式：{file_ext}。请使用Excel(.xlsx/.xls)或CSV(.csv)格式', 'danger')
                    return redirect(request.url)
                
                # 检查文件是否为空
                if df.empty:
                    flash('文件内容为空，请检查文件是否正确', 'danger')
                    return redirect(request.url)
                
                # 检查必要的列是否存在
                required_cols = ['*微信名', '*手机号', '*订单编码', '*订单信息', '*订单类型', '*完成时间', '*数量']
                missing_cols = []
                for col in required_cols:
                    if col not in df.columns and col.replace('*', '') not in df.columns:
                        missing_cols.append(col)
                
                if missing_cols:
                    flash(f'文件缺少必要的列：{", ".join(missing_cols)}。请检查模板格式', 'danger')
                    return redirect(request.url)
                
                success_count = 0
                error_count = 0
                errors = []
                
                # 获取订单类型映射
                order_types = {ot.name: ot.id for ot in OrderType.query.all()}
                
                for index, row in df.iterrows():
                    try:
                        # 检查必填字段（支持带星号的列名）
                        phone_col = '*手机号' if '*手机号' in df.columns else '手机号'
                        wechat_col = '*微信名' if '*微信名' in df.columns else '微信名'
                        order_code_col = '*订单编码' if '*订单编码' in df.columns else '订单编码'
                        order_info_col = '*订单信息' if '*订单信息' in df.columns else '订单信息'
                        order_type_col = '*订单类型' if '*订单类型' in df.columns else '订单类型'
                        completion_time_col = '*完成时间' if '*完成时间' in df.columns else '完成时间'
                        quantity_col = '*数量' if '*数量' in df.columns else '数量'
                        
                        if pd.isna(row.get(phone_col)) or str(row.get(phone_col)).strip() == '':
                            errors.append(f"第{index+2}行：手机号不能为空")
                            error_count += 1
                            continue
                        
                        if pd.isna(row.get(wechat_col)) or str(row.get(wechat_col)).strip() == '':
                            errors.append(f"第{index+2}行：微信名不能为空")
                            error_count += 1
                            continue
                        
                        if pd.isna(row.get(order_code_col)) or str(row.get(order_code_col)).strip() == '':
                            errors.append(f"第{index+2}行：订单编码不能为空")
                            error_count += 1
                            continue
                        
                        if pd.isna(row.get(order_info_col)) or str(row.get(order_info_col)).strip() == '':
                            errors.append(f"第{index+2}行：订单信息不能为空")
                            error_count += 1
                            continue
                        
                        if pd.isna(row.get(order_type_col)) or str(row.get(order_type_col)).strip() == '':
                            errors.append(f"第{index+2}行：订单类型不能为空")
                            error_count += 1
                            continue
                        
                        if pd.isna(row.get(completion_time_col)) or str(row.get(completion_time_col)).strip() == '':
                            errors.append(f"第{index+2}行：完成时间不能为空")
                            error_count += 1
                            continue
                        
                        if pd.isna(row.get(quantity_col)) or str(row.get(quantity_col)).strip() == '':
                            errors.append(f"第{index+2}行：数量不能为空")
                            error_count += 1
                            continue
                        
                        # 处理完成时间
                        completion_time = None
                        if not pd.isna(row.get(completion_time_col)):
                            try:
                                if isinstance(row[completion_time_col], str):
                                    completion_time = datetime.strptime(row[completion_time_col], '%Y-%m-%d')
                                else:
                                    completion_time = pd.to_datetime(row[completion_time_col]).to_pydatetime()
                            except:
                                pass
                        
                        # 处理订单类型
                        order_type_id = None
                        if not pd.isna(row.get(order_type_col)):
                            order_type_name = str(row[order_type_col]).strip()
                            order_type_id = order_types.get(order_type_name)
                        
                        # 创建订单（使用正确的列名，处理NaN值）
                        order = Order(
                            order_code=str(row.get(order_code_col, '')).strip() if not pd.isna(row.get(order_code_col)) else '',
                            wechat_name=str(row.get(wechat_col, '')).strip() if not pd.isna(row.get(wechat_col)) else '',
                            wechat_id=str(row.get('微信号', '')).strip() if not pd.isna(row.get('微信号')) else '',
                            phone=str(row.get(phone_col, '')).strip() if not pd.isna(row.get(phone_col)) else '',
                            order_info=str(row.get(order_info_col, '')).strip() if not pd.isna(row.get(order_info_col)) else '',
                            completion_time=completion_time,
                            quantity=int(row.get(quantity_col, 0)) if not pd.isna(row.get(quantity_col)) else None,
                            amount=float(row.get('金额', 0)) if not pd.isna(row.get('金额')) else None,
                            notes=str(row.get('备注', '')).strip() if not pd.isna(row.get('备注')) else '',
                            user_id=current_user.id,
                            order_type_id=order_type_id,
                            status=str(row.get('状态', '未完成')).strip() if not pd.isna(row.get('状态')) else '未完成'
                        )
                        
                        db.session.add(order)
                        success_count += 1
                        
                    except Exception as e:
                        errors.append(f"第{index+1}行：{str(e)}")
                        error_count += 1
                
                db.session.commit()
                
                flash(f'导入完成！成功：{success_count}条，失败：{error_count}条', 'success')
                
                if errors:
                    flash('错误详情：' + '; '.join(errors[:10]), 'warning')
                    if len(errors) > 10:
                        flash(f'...还有{len(errors)-10}个错误', 'warning')
                
                return redirect(url_for('main.order_list'))
                
            except pd.errors.EmptyDataError:
                flash('文件内容为空或格式不正确，请检查文件', 'danger')
                return redirect(request.url)
            except pd.errors.ParserError as e:
                flash(f'文件解析失败：{str(e)}。请检查文件格式是否正确', 'danger')
                return redirect(request.url)
            except UnicodeDecodeError:
                flash('文件编码错误，请保存为UTF-8编码的CSV文件或使用Excel格式', 'danger')
                return redirect(request.url)
            except FileNotFoundError:
                flash('文件未找到，请重新选择文件', 'danger')
                return redirect(request.url)
            except PermissionError:
                flash('文件被占用，请关闭文件后重试', 'danger')
                return redirect(request.url)
            except Exception as e:
                error_msg = str(e)
                if 'Excel file format cannot be determined' in error_msg:
                    flash('Excel文件格式无法识别，请确保文件未损坏', 'danger')
                elif 'No such file or directory' in error_msg:
                    flash('文件路径错误或文件不存在', 'danger')
                elif 'BadZipFile' in error_msg:
                    flash('Excel文件已损坏，请重新保存文件', 'danger')
                else:
                    flash(f'文件处理失败：{error_msg}', 'danger')
                return redirect(request.url)
        else:
            flash('不支持的文件格式', 'danger')
            return redirect(request.url)
    
    return render_template('main/import_orders.html')

@main.route('/order/image/delete/<int:id>', methods=['POST'])
@login_required
def delete_image(id):
    """删除订单图片"""
    image = OrderImage.query.get_or_404(id)
    order = image.order
    
    # 权限检查
    if not current_user.can(Permission.VIEW_ALL) and order.user_id != current_user.id:
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({'success': False, 'error': '权限不足'})
        abort(403)
    
    try:
        # 删除文件
        image_path = os.path.join(current_app.config['UPLOAD_FOLDER'], image.image_path.replace('uploads/', ''))
        if os.path.exists(image_path):
            os.remove(image_path)
        
        db.session.delete(image)
        db.session.commit()
        
        # 根据请求类型返回不同响应
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({'success': True, 'message': '图片删除成功！'})
        else:
            flash('图片删除成功！', 'success')
            return redirect(url_for('main.edit_order', id=order.id))
            
    except Exception as e:
        db.session.rollback()
        print(f"删除图片失败: {e}")
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({'success': False, 'error': f'删除失败：{str(e)}'})
        else:
            flash(f'删除失败：{str(e)}', 'error')
            return redirect(url_for('main.edit_order', id=order.id))

@main.route('/quick_add', methods=['POST'])
@login_required
def quick_add_order():
    """快速添加订单的API"""
    try:
        # 生成订单编号
        order_code = f"ORD{datetime.now().strftime('%Y%m%d%H%M%S')}{uuid.uuid4().hex[:8].upper()}"
        
        # 创建订单
        order = Order(
            order_code=order_code,
            wechat_name=request.form.get('wechat_name', '').strip(),
            wechat_id=request.form.get('wechat_id', '').strip(),
            phone=request.form.get('phone', '').strip(),
            order_info=request.form.get('order_info', '').strip(),
            quantity=request.form.get('quantity', type=int),
            amount=request.form.get('amount', type=float),
            notes=request.form.get('notes', '').strip(),
            user_id=current_user.id,
            order_type_id=request.form.get('order_type_id', type=int)
        )
        
        # 处理完成时间
        completion_time_str = request.form.get('completion_time')
        if completion_time_str:
            try:
                order.completion_time = datetime.strptime(completion_time_str, '%Y-%m-%d')
            except ValueError:
                return jsonify({'success': False, 'message': '完成时间格式错误'})
        
        db.session.add(order)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': '订单创建成功！',
            'order_id': order.id,
            'order_code': order.order_code
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'创建失败：{str(e)}'})

@main.route('/uploads/<path:filename>')
@login_required
def uploaded_file(filename):
    """提供上传文件访问"""
    # 安全检查：防止路径遍历攻击
    if '..' in filename or filename.startswith('/'):
        abort(404)
    
    # 检查文件是否存在
    file_path = os.path.join(current_app.config['UPLOAD_FOLDER'], filename)
    if not os.path.exists(file_path):
        abort(404)
    
    # 验证文件扩展名
    if not allowed_file(filename):
        abort(403)
    
    return send_from_directory(current_app.config['UPLOAD_FOLDER'], filename)

@main.route('/order/update_status/<int:order_id>', methods=['POST'])
@login_required
def update_order_status(order_id):
    """更新订单状态"""
    order = Order.query.get_or_404(order_id)
    
    # 权限检查
    if not current_user.can(Permission.VIEW_ALL) and order.user_id != current_user.id:
        return jsonify({'success': False, 'error': '权限不足'})
    
    try:
        # 支持JSON和表单数据
        if request.is_json:
            data = request.get_json()
            new_status = data.get('status')
        else:
            new_status = request.form.get('status')
        
        if not new_status:
            return jsonify({'success': False, 'error': '状态不能为空'})
        
        # 验证状态值
        valid_statuses = ['未完成', '已完成', '已结算', '未结算']
        if new_status not in valid_statuses:
            return jsonify({'success': False, 'error': '无效的状态值'})
        
        order.status = new_status
        
        # 如果状态改为已完成，自动设置完成时间
        if new_status == '已完成' and not order.completion_time:
            order.completion_time = datetime.now()
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': '状态更新成功',
            'new_status': new_status,
            'completion_time': order.completion_time.strftime('%Y-%m-%d') if order.completion_time else None
        })
        
    except Exception as e:
        db.session.rollback()
        print(f"更新订单状态失败: {e}")
        return jsonify({'success': False, 'error': f'更新失败：{str(e)}'})

@main.route('/batch_update_status', methods=['POST'])
@login_required
def batch_update_status():
    """批量更新订单状态"""
    if not current_user.can(Permission.VIEW_ALL):
        return jsonify({'success': False, 'error': '权限不足'})
    
    try:
        # 支持JSON和表单数据
        if request.is_json:
            data = request.get_json()
            order_ids = data.get('order_ids', [])
            new_status = data.get('status')
        else:
            order_ids = request.form.getlist('order_ids[]')
            new_status = request.form.get('status')
        
        if not order_ids:
            return jsonify({'success': False, 'error': '请选择要更新的订单'})
        
        if not new_status:
            return jsonify({'success': False, 'error': '状态不能为空'})
        
        # 验证状态值
        valid_statuses = ['未完成', '已完成', '已结算', '未结算']
        if new_status not in valid_statuses:
            return jsonify({'success': False, 'error': '无效的状态值'})
        
        success_count = 0
        error_count = 0
        
        for order_id in order_ids:
            try:
                order = Order.query.get(int(order_id))
                if order:
                    order.status = new_status
                    
                    # 如果状态改为已完成，自动设置完成时间
                    if new_status == '已完成' and not order.completion_time:
                        order.completion_time = datetime.now()
                    
                    success_count += 1
                else:
                    error_count += 1
            except Exception as e:
                error_count += 1
                print(f"更新订单 {order_id} 失败: {e}")
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': f'批量更新完成！成功：{success_count}条，失败：{error_count}条'
        })
        
    except Exception as e:
        db.session.rollback()
        print(f"批量更新状态失败: {e}")
        return jsonify({'success': False, 'error': f'批量更新失败：{str(e)}'})

@main.route('/batch_delete_orders', methods=['POST'])
@login_required
def batch_delete_orders():
    """批量删除订单"""
    if not current_user.can(Permission.VIEW_ALL):
        return jsonify({'success': False, 'error': '权限不足'})
    
    try:
        # 支持JSON和表单数据
        if request.is_json:
            data = request.get_json()
            order_ids = data.get('order_ids', [])
        else:
            order_ids = request.form.getlist('order_ids[]')
        
        if not order_ids:
            return jsonify({'success': False, 'error': '请选择要删除的订单'})
        
        success_count = 0
        error_count = 0
        
        for order_id in order_ids:
            try:
                order = Order.query.get(int(order_id))
                if order:
                    # 删除相关图片
                    for image in order.images:
                        try:
                            image_path = os.path.join(current_app.config['UPLOAD_FOLDER'], image.image_path.replace('uploads/', ''))
                            if os.path.exists(image_path):
                                os.remove(image_path)
                        except Exception as e:
                            print(f"删除图片失败: {e}")
                            pass
                        db.session.delete(image)
                    
                    db.session.delete(order)
                    success_count += 1
                else:
                    error_count += 1
            except Exception as e:
                error_count += 1
                print(f"删除订单 {order_id} 失败: {e}")
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': f'批量删除完成！成功：{success_count}条，失败：{error_count}条'
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'批量删除失败：{str(e)}'})

@main.route('/backup/database', methods=['POST'])
@login_required
@admin_required
def backup_database():
    """备份数据库"""
    try:
        import shutil
        from datetime import datetime
        
        # 数据库文件路径
        db_path = current_app.config['SQLALCHEMY_DATABASE_URI'].replace('sqlite:///', '')
        if not os.path.exists(db_path):
            return jsonify({'success': False, 'message': '数据库文件不存在'})
        
        # 创建备份目录
        backup_dir = os.path.join(current_app.root_path, '..', 'backups')
        os.makedirs(backup_dir, exist_ok=True)
        
        # 生成备份文件名
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        backup_filename = f'data_backup_{timestamp}.sqlite'
        backup_path = os.path.join(backup_dir, backup_filename)
        
        # 复制数据库文件
        shutil.copy2(db_path, backup_path)
        
        return jsonify({
            'success': True,
            'message': f'数据库备份成功！备份文件：{backup_filename}',
            'backup_path': backup_path
        })
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'备份失败：{str(e)}'})